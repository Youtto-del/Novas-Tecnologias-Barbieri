import pandas as pd
from openpyxl import load_workbook

# Carrega as tabelas
lista_de_notas = 'Lista de notas.xlsx'

# Importando e limpando planilha de desdobramentos
desdobramentos = pd.read_excel('relatorio_desdobramentos.xlsx')
colunas_tabela = list(desdobramentos.columns)
colunas_novas = list(desdobramentos.iloc[2])
dici = {}
for x in range(len(colunas_tabela)):
  dici[colunas_tabela[x]] = colunas_novas[x]
desdobramentos.rename(columns=dici, inplace=True)
desdobramentos.drop(axis=0, index=[0, 1, 2], inplace=True)

# Contrói dataframe da lista de notas
df = pd.read_excel(lista_de_notas)

# filtra os casos de "Migrado" e "Digitalizado"
resultado_linha = []
erros = []
for linha in df.iterrows():
    if linha[1]['Status 1'] == 'Migrado' or linha[1]['Status 1'] == 'Digitalizado':
        insercao = [linha[1]['Processo'], linha[1]['originario_1'], linha[1]['Status 1']]
        resultado_linha.append(insercao)
    elif linha[1]['Status 2'] == 'Migrado' or linha[1]['Status 2'] == 'Digitalizado':
        insercao = [linha[1]['Processo'], linha[1]['originario_2'], linha[1]['Status 2']]
        resultado_linha.append(insercao)
    elif linha[1]['Status 2'] == 'Migrado' or linha[1]['Status 3'] == 'Digitalizado':
        insercao = [linha[1]['Processo'], linha[1]['originario_3'], linha[1]['Status 3']]
        resultado_linha.append(insercao)
    else:
        erros.append(linha[1])

# cria e exportar dataframe do resultado filtrado
df_filtrado = pd.DataFrame(resultado_linha, columns=['Processo', 'Originario', 'Status'])
df_filtrado.to_excel('Resultado filtrado.xlsx', index=False)
print(df_filtrado)
print('Processos fora das hipóteses de digitalização:', len(erros))

teste = [y for y in df_filtrado['Processo'] if y in list(desdobramentos['numero'])]
print(teste)

resultado_final = []
for item in df_filtrado.iterrows():
    if item[1]['Originario'] in list(desdobramentos['numero']):
        index = list(desdobramentos['numero']).index(item[1]['Originario'])
        pasta_desdobramento = desdobramentos.iloc[index, 2]
        print(pasta_desdobramento)
        resultado_final.append([pasta_desdobramento[:9],
                                pasta_desdobramento,
                                item[1]['Originario'],
                                item[1]['Processo']])
    else:
        print(item[1]['Originario'], 'não encontrado')

print(resultado_final)


# Crie um dataframe com os novos dados
df_resultado_final = pd.DataFrame(resultado_final, columns=['Pasta', 'Pasta desdobramento', 'Número antigo', 'Processo'])

# Carregue a planilha existente
modelo_att = load_workbook('Modelo EPROC ATT.xlsx')
worksheet = modelo_att.active

# Selecione a planilha que você deseja adicionar o novo dataframe
writer = pd.ExcelWriter('Modelo EPROC ATT.xlsx', engine='openpyxl')
writer.book = modelo_att

# Adicione o novo dataframe na planilha existente
df_resultado_final.to_excel(writer, index=False, header=False, sheet_name='Importacao', startrow=worksheet.max_row)

# Salve as mudanças na planilha
writer.close()

